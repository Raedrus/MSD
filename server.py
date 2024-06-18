from flask import Flask, request, jsonify
import pandas as pd
import os
import traceback
from datetime import datetime, timedelta
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')  # Set the backend to 'Agg' to avoid tkinter issues
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from io import BytesIO
import time
import matplotlib.dates as mdates

# pip install flask openpyxl matplotlib

app = Flask(__name__)

# Excel file path (adjust to your specific path)
excel_file = r'C:\Users\hueyu\OneDrive\Documents\rubbish_data.xlsx'

# Initialize reset_count and reset_count_per_hour
reset_count = 0
reset_count_per_hour = 0
last_hour = None

# Check if Excel file exists, load existing data if it does
if os.path.exists(excel_file):
    df = pd.read_excel(excel_file, sheet_name='Sheet1')  # Adjust sheet name as needed

    # Read the latest reset_count and reset_count_per_hour from the Excel file
    if not df.empty and 'Reset_Count_per_day' in df.columns:
        reset_count = df['Reset_Count_per_day'].iloc[-1]  # Get the last value in the column
    if not df.empty and 'Reset_Count_per_hour' in df.columns:
        reset_count_per_hour = df['Reset_Count_per_hour'].iloc[-1]  # Get the last value in the column
else:
    columns = ['time', 'E_device', 'General Waste', 'Battery', 'Alu_can', 'Reset_Count_per_hour', 'Reset_Count_per_day']
    df = pd.DataFrame(columns=columns)

current_date = datetime.now().date()

def check_reset_count_per_hour(timestamp, last_timestamp):
    # Reset count per hour if the new data is in a different hour
    if last_timestamp and last_timestamp.strftime('%Y-%m-%d %H') != timestamp.strftime('%Y-%m-%d %H'):
        return 0
    return reset_count_per_hour

def load_existing_data():
    global df, current_date, reset_count, reset_count_per_hour
    try:
        df = pd.read_excel(excel_file)
        if not df.empty:
            last_entry_date = df['time'].max()
            last_entry_date = datetime.strptime(last_entry_date, '%Y-%m-%d %H:%M:%S')
            if last_entry_date.date() != datetime.now().date():
                reset_count = 0
                reset_count_per_hour = 0
            current_date = datetime.now().date()
    except FileNotFoundError:
        df = pd.DataFrame()
    except Exception as e:
        print("Error loading existing data:", e)
        print(traceback.format_exc())

# def generate_charts(df):
#     try:
#         # Ensure the necessary columns are numeric
#         df['E_device'] = pd.to_numeric(df['E_device'], errors='coerce')
#         df['General Waste'] = pd.to_numeric(df['General Waste'], errors='coerce')
#         df['Battery'] = pd.to_numeric(df['Battery'], errors='coerce')
#         df['Alu_can'] = pd.to_numeric(df['Alu_can'], errors='coerce')
#         df['Reset_Count_per_hour'] = pd.to_numeric(df['Reset_Count_per_hour'], errors='coerce')
#         df['Reset_Count_per_day'] = pd.to_numeric(df['Reset_Count_per_day'], errors='coerce')
#
#         # Drop rows with non-numeric values
#         df.dropna(subset=['E_device', 'General Waste', 'Battery', 'Alu_can', 'Reset_Count_per_hour', 'Reset_Count_per_day'], inplace=True)
#
#         # Create a figure and a set of subplots
#         fig, axs = plt.subplots(3, 1, figsize=(10, 15))
#
#         # First chart: Total rubbish quantity over time (stacked bar chart)
#         df['time'] = pd.to_datetime(df['time'])
#         df.set_index('time', inplace=True)
#         df[['E_device', 'General Waste', 'Battery', 'Alu_can']].plot(kind='bar', stacked=True, ax=axs[0])
#         axs[0].set_title('Total Rubbish Quantity Over Time')
#         axs[0].set_xlabel('Time')
#         axs[0].set_ylabel('Quantity')
#         axs[0].xaxis.set_major_locator(mdates.HourLocator(interval=6))
#         axs[0].xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d %H:%M'))
#         axs[0].legend(loc='upper left')
#
#         # Second chart: Reset count per hour for the past 30 hours
#         past_30_hours = datetime.now() - timedelta(hours=30)
#         df_30_hours = df[df.index >= past_30_hours]
#         df_30_hours_hourly = df_30_hours.resample('H').max()
#         df_30_hours_hourly['Reset_Count_per_hour'].plot(kind='bar', ax=axs[1])
#         axs[1].set_title('Reset Count Per Hour (Past 30 Hours)')
#         axs[1].set_xlabel('Hour')
#         axs[1].set_ylabel('Reset Count')
#         axs[1].xaxis.set_major_locator(mdates.HourLocator(interval=1))
#         axs[1].xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d %H:%M'))
#
#         # Third chart: Reset count per day for the past 2 weeks
#         past_2_weeks = datetime.now() - timedelta(days=14)
#         df_2_weeks = df[df.index >= past_2_weeks]
#         df_2_weeks_daily = df_2_weeks.resample('D').max()
#         df_2_weeks_daily['Reset_Count_per_day'].plot(kind='bar', ax=axs[2])
#         axs[2].set_title('Reset Count Per Day (Past 2 Weeks)')
#         axs[2].set_xlabel('Day')
#         axs[2].set_ylabel('Reset Count')
#         axs[2].xaxis.set_major_locator(mdates.DayLocator(interval=1))
#         axs[2].xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
#
#         plt.tight_layout()
#
#         # Save the plot to a BytesIO object
#         buf = BytesIO()
#         plt.savefig(buf, format='png')
#         buf.seek(0)
#
#         # Load the Excel file
#         workbook = load_workbook(excel_file)
#         sheet = workbook.active
#
#         # Insert the chart images into the Excel file
#         for i, ax in enumerate(axs):
#             img = Image(buf)
#             img.anchor = f'A{30 * (i + 1)}'
#             sheet.add_image(img)
#
#         workbook.save(excel_file)
#     except Exception as e:
#         print("Error generating charts:", e)
#         print(traceback.format_exc())

def generate_and_embed_charts():
    try:
        # Load existing data from Excel file
        excel_file = r'C:\Users\hueyu\OneDrive\Documents\rubbish_data.xlsx'  # Replace with your actual Excel file path
        if os.path.exists(excel_file):
            df_from_excel = pd.read_excel(excel_file, sheet_name='Sheet1')
        else:
            df_from_excel = pd.DataFrame(
                columns=['time', 'E_device', 'General Waste', 'Battery', 'Alu_can', 'Reset_Count_per_hour',
                         'Reset_Count_per_day'])

        # Concatenate existing data from Excel with current in-memory DataFrame (df)
        df_combined = pd.concat([df_from_excel, df], ignore_index=True)

        # Create lists to store data for plotting
        time_points = []
        e_device_values = []
        general_waste_values = []
        battery_values = []
        alu_can_values = []

        for index, row in df.iterrows():
            time_points.append(row['time'])
            e_device_values.append(row['E_device'])
            general_waste_values.append(row['General Waste'])
            battery_values.append(row['Battery'])
            alu_can_values.append(row['Alu_can'])

        plt.figure(figsize=(20, 8))

        plt.bar(time_points, e_device_values, label='E_device', color='blue')
        plt.bar(time_points, general_waste_values, bottom=e_device_values, label='General Waste', color='green')
        plt.bar(time_points, battery_values, bottom=[sum(x) for x in zip(e_device_values, general_waste_values)],
                label='Battery', color='orange')
        plt.bar(time_points, alu_can_values,
                bottom=[sum(x) for x in zip(e_device_values, general_waste_values, battery_values)],
                label='Alu_can', color='red')

        plt.xlabel('Time')
        plt.ylabel('Quantity')
        plt.title('Total Rubbish Quantities Over Time')
        plt.legend()
        plt.xticks(rotation=45)
        plt.tight_layout()

        # Save the plot to a temporary buffer
        image_stream = BytesIO()
        plt.savefig(image_stream, format='png')
        image_stream.seek(0)

        # Open the Excel workbook
        wb = load_workbook(excel_file)
        ws = wb.active

        # Add the image to Excel sheet (starting from column H)
        img = Image(image_stream)
        ws.add_image(img, 'J3')  # Adjust the cell reference as needed

        # Save the Excel file
        wb.save(excel_file)

        print("Chart embedded successfully.")

        # Concatenate existing data from Excel with current in-memory DataFrame (df)
        df_combined = pd.concat([df_from_excel, df], ignore_index=True)

        # Ensure datetime column is in datetime format
        df_combined['time'] = pd.to_datetime(df_combined['time'])

        # Extracting reset count per hour data
        df_reset_count = df_combined[['time', 'Reset_Count_per_hour']].copy()

        # Group by hour and find maximum reset count for each hour
        df_reset_count['hour'] = df_reset_count['time'].dt.hour
        max_reset_count_per_hour = df_reset_count.groupby('hour')['Reset_Count_per_hour'].max()

        # Plotting the chart
        plt.figure(figsize=(12, 6))
        plt.bar(max_reset_count_per_hour.index, max_reset_count_per_hour.values, color='blue')

        plt.xlabel('Hour of Day')
        plt.ylabel('Highest Reset Count per Hour')
        plt.title('Highest Reset Count per Hour')
        plt.xticks(range(24))
        plt.tight_layout()

        # Save the plot to a temporary buffer
        image_stream = BytesIO()
        plt.savefig(image_stream, format='png')
        image_stream.seek(0)

        # Open the Excel workbook
        wb = load_workbook(excel_file)
        ws = wb.active

        # Add the image to Excel sheet (starting from column H)
        img = Image(image_stream)
        ws.add_image(img, 'J47')  # Adjust the cell reference as needed

        # Save the Excel file
        wb.save(excel_file)

        print("Chart embedded successfully.")

    except Exception as e:
        print("Error occurred while generating or embedding charts:", e)


@app.route('/submit_data', methods=['POST'])
def submit_data():
    global df, reset_count, reset_count_per_hour, current_date
    try:
        data = request.json
        if not data:
            return jsonify({"error": "No data provided"}), 400

        timestamp_str = data.get('time')
        timestamp = datetime.strptime(timestamp_str, '%Y-%m-%d %H:%M:%S')
        e_device_quantity = data.get('E_device', 0)
        general_waste_quantity = data.get('General Waste', 0)
        battery_quantity = data.get('Battery', 0)
        alu_can_quantity = data.get('Alu_can', 0)
        dustbin_reset = data.get('Reset_Count_per_day', False)
        event = data.get('Event', '')

        # Load existing data and check if it's a new day
        load_existing_data()

        if not df.empty:
            last_entry_time = datetime.strptime(df['time'].max(), '%Y-%m-%d %H:%M:%S')
        else:
            last_entry_time = None

        reset_count_per_hour = check_reset_count_per_hour(timestamp, last_entry_time)

        if timestamp.date() != current_date:
            reset_count = 0
            reset_count_per_hour = 0
            current_date = timestamp.date()

        if last_entry_time and last_entry_time.hour != timestamp.hour:
            reset_count_per_hour = 0

        if dustbin_reset:
            reset_count += 1
            reset_count_per_hour += 1

            new_row = {
                'time': [timestamp_str],
                'E_device': [0],
                'General Waste': [0],
                'Battery': [0],
                'Alu_can': [0],
                'Reset_Count_per_hour': [reset_count_per_hour],
                'Reset_Count_per_day': [reset_count],
                'Event': [event]
            }
        else:
            new_row = {
                'time': [timestamp_str],
                'E_device': [e_device_quantity],
                'General Waste': [general_waste_quantity],
                'Battery': [battery_quantity],
                'Alu_can': [alu_can_quantity],
                'Reset_Count_per_hour': [reset_count_per_hour],
                'Reset_Count_per_day': [reset_count],
                'Event': [event]
            }

        new_data = pd.DataFrame(new_row)
        df = pd.concat([df, new_data], ignore_index=True)

        # Save DataFrame to Excel file
        df.to_excel(excel_file, index=False)

        # Generate and embed charts
        generate_and_embed_charts()

        # Ensure OneDrive sync
        time.sleep(3)

        return jsonify({"message": "Data submitted successfully"}), 200
    except Exception as e:
        print("Error processing data:", e)
        print(traceback.format_exc())
        return jsonify({"error": "An error occurred while processing the data"}), 500

if __name__ == '__main__':
    app.run(debug=True)
    #generate_and_embed_charts()
